import openpyxl
import pandas as pd
from utils import suspicious_check_function
from openpyxl.styles import PatternFill

class DataAnalyzer:
    def analyze_data(self, dataframes):
        # Initialize dictionaries to store results
        products_per_supplier = {}
        total_value_per_supplier = {}
        products_under_10_inv = {}

        for df in dataframes:
            for index, row in df.iterrows():
                supplier_name = row['SupplierName']
                inventory = row['Inventory']
                price = row['Price']
                product_num = row['ProductNumber']

                # Calculation of number of products per supplier
                if supplier_name in products_per_supplier:
                    current_num_products = products_per_supplier.get(supplier_name)
                    products_per_supplier[supplier_name] = current_num_products + 1
                else:
                    products_per_supplier[supplier_name] = 1

                # Calculation of total value of inventory per supplier
                if supplier_name in total_value_per_supplier:
                    current_total_value = total_value_per_supplier.get(supplier_name)
                    total_value_per_supplier[supplier_name] = current_total_value + inventory * price
                else:
                    total_value_per_supplier[supplier_name] = inventory * price

                # Check for suspicious values
                if suspicious_check_function(inventory):
                    products_under_10_inv[int(product_num)] = int(inventory)

        return {
            'products_per_supplier': products_per_supplier,
            'total_value_per_supplier': total_value_per_supplier,
            'products_under_10_inv': products_under_10_inv
        }

    def print_results(self, results):
        print("Products per Supplier:")
        print(results['products_per_supplier'])

        print("Total Value per Supplier:")
        print(results['total_value_per_supplier'])

        print("Products with Inventory < 10:")
        print(results['products_under_10_inv'])

    def save_results(self, results, output_filename):
        # Save results to an Excel file
        inv_file = openpyxl.Workbook()
        product_list = inv_file.active

        for supplier, num_products in results['products_per_supplier'].items():
            product_list.append([supplier, num_products])

        for supplier, total_value in results['total_value_per_supplier'].items():
            product_list.append([supplier, total_value])

        for product_num, inventory in results['products_under_10_inv'].items():
            product_list.append([product_num, inventory])

        inv_file.save(output_filename)

    def save_highlighted_results(self, dataframes, output_filename):
        # Create a workbook for saving highlighted results
        result_file = openpyxl.Workbook()
        result_sheet = result_file.active

        # Define a fill style for highlighting suspicious values
        suspicious_fill = PatternFill(start_color="FFFF00", end_color="FF0000", fill_type="solid")

        row_index = 1  # Initialize the row index

        for df in dataframes:
            for index, row in df.iterrows():
                inventory = row['Inventory']

                # Check for suspicious values
                if suspicious_check_function(inventory):
                    # Highlight the row with suspicious values
                    for cell in result_sheet[row_index]:
                        cell.fill = suspicious_fill

                # Add your specific analysis code here
                # Example: Calculate additional metrics, perform data transformations, etc.
                
                # Write the data to the result sheet
                result_sheet.append([row['SupplierName'], inventory, row['Price'], row['ProductNumber']])

                row_index += 1  # Increment the row index

        result_file.save(output_filename)
